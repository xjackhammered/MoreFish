import logging
import os
import sys
import time
from datetime import datetime, timedelta
from decimal import Decimal
import traceback
from django.http import FileResponse, HttpRequest, HttpResponse, JsonResponse
import uuid
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import requests as requests
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import json
# Create your views here.
from pytz import timezone
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
# from device.forms import RawDataCalulationAdminForm, convert_ph_value, convert_tds_value, convert_temperature_value
from device.helper import (
    pond_wise_devices_daily_data,
    pond_wise_devices_monthly_data,
    pond_wise_devices_sensor_monthly_data,
    pond_wise_devices_sensor_yearly_data,
    pond_wise_devices_sensors_daily_data,
    pond_wise_devices_sensors_weekly_data,
    pond_wise_devices_weekly_data,
    pond_wise_devices_yearly_data,
)
import device.publisher as publisher
from assets.serializers import AssetsSerializer
from assets.models import AssetsProperties, AssetsFiles
from device.models import (
    Complain,
    ComplainCategory,
    DeviceGateway,
    Device,
    DeviceData,
    DeviceControlInfo,
    Camera,
    InvalidValue,
    SensorConfiguration,
    UserManualData,
    Weather,
)
from asgiref.sync import async_to_sync
from device.serializers import (

    AssetsCameraSerializer,
    PondDataSerializer,
    ResultDataSerializer,
    WeatherSerializer,
    UserManualDataSerializer
)
from django.db.models import Avg
import paho.mqtt.client as mqtt
from device.service import DeviceService
from morefish_pppl import settings
from helper import unique
from notification.models import Notifications
from users.models import Company, Phone, User
# from weather_report import get_weather_report, check_last_weather_report

from django.db.models import Prefetch
from django.db.models import Q
from drf_spectacular.utils import extend_schema

from .mqtt_utils import publish_command


def is_admin(user):
    return user.groups.filter(name="admin").exists()



class AssetsListWithData(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request: HttpRequest):
        try:
            company_id = request.GET.get('company_id',None)
                
            result_data =  DeviceService.dashboard_data(user=request.user,company_id=company_id)
            
            serialized_data = ResultDataSerializer(data={"result_data": result_data})

            serialized_data.is_valid(raise_exception=True)                
            return Response({
                        "success": "True",
                        "status_code": status.HTTP_200_OK,
                        "message": "Asset List",
                        "data": serialized_data.data,
                    })
        except Exception as e:
            traceback.print_exc()
            response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
            return Response(response, status=status.HTTP_201_CREATED)

class PondList(APIView):
    permission_classes = (IsAuthenticated,)
    
    def get(self,request:HttpRequest):
        ponds = DeviceService.get_asset_list(user=request.user,company_id=request.user.company_id)
        
        return Response({
                        "success": "True",
                        "status_code": status.HTTP_200_OK,
                        "message": "Asset List",
                        "data": ponds,
                    })        

#
# class PondData(APIView):
#     permission_classes = (IsAuthenticated,)
#     serializer_class = ResultDataSerializer
#
#     def get(self,request:HttpRequest):
#         """
#         Retrieve data for a pond.
#
#         Parameters:
#         - name: asset_id
#           description: The ID of the asset
#           required: true
#           type: integer
#           in: query
#         """
#         serializer = PondDataSerializer(data=request.query_params)
#         serializer.is_valid(raise_exception=True)
#         asset_id = serializer.validated_data["asset_id"]
#         service_start_time = time.time()
#         asset_data = DeviceService.get_asset_data(asset_id=asset_id,user=request.user,company_id=request.user.company_id)
#         service_end_time = time.time()
#         serialized_data = ResultDataSerializer(data={"result_data": asset_data})
#         start_serialization = time.time()
#         service_duration = service_end_time - service_start_time
#         print(f"Service Duration: {service_duration} seconds")
#
#         try:
#             serialized_data.is_valid(raise_exception=True)
#             end_serialization = time.time()
#             serialization_duration = end_serialization - start_serialization
#             print(f"Serialization Duration: {serialization_duration} seconds")
#             if len(asset_data.get("devices"))==0:
#
#                 return Response({
#                     "success": "True",
#                     "status_code": status.HTTP_404_NOT_FOUND,
#                     "message": "This pond does not have any device installed",
#                     "data": serialized_data.data,
#                 })
#             return Response({
#                 "success": "True",
#                 "status_code": status.HTTP_200_OK,
#                 "message": "Asset List",
#                 "data": serialized_data.data,
#             })
#         except Exception as e:
#             traceback.print_exc()
#             response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
#             return Response(response, status=status.HTTP_500_INTERNAL_SERVER_ERROR)






class PondData(APIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = ResultDataSerializer

    def get(self, request: HttpRequest):
        """
        Retrieve data for a pond, including its devices and each device’s aerators.
        Query params:
        - asset_id (int, required)
        """
        # 1) validate query params
        qs = PondDataSerializer(data=request.query_params)
        qs.is_valid(raise_exception=True)
        asset_id = qs.validated_data["asset_id"]

        # 2) fetch asset_data, catch missing asset
        try:
            t0 = time.time()
            asset_data = DeviceService.get_asset_data(
                asset_id=asset_id,
                user=request.user,
                company_id=request.user.company_id
            )
            t1 = time.time()
            print(f"Service Duration: {t1 - t0:.3f}s")
        except AssetsProperties.DoesNotExist:
            return Response({
                "success": "False",
                "status_code": status.HTTP_404_NOT_FOUND,
                "message": f"No asset found with id={asset_id}",
                "data": {}
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            traceback.print_exc()
            return Response({
                "success": "False",
                "status_code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"Error retrieving asset data: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 3) inject aerators into each device
        for device in asset_data.get("devices", []):
            dev_pk = device.get("device_id")  # ← use "device_id" here
            if not dev_pk:
                continue
            qs = Aerator.objects.filter(device_id=dev_pk, is_active=True)
            # serialize them instead of building by hand
            device["aerators"] = AeratorInfoSerializer(qs, many=True).data

        # 4) serialize & unwrap
        t2 = time.time()
        ser = ResultDataSerializer(data={"result_data": asset_data})
        ser.is_valid(raise_exception=True)
        payload = ser.data["result_data"]
        t3 = time.time()
        print(f"Serialization Duration: {t3 - t2:.3f}s")

        # 5) final response
        devices = payload.get("devices", [])
        if not devices:
            return Response({
                "success": "True",
                "status_code": status.HTTP_404_NOT_FOUND,
                "message": "This pond has no devices.",
                "data": payload,
            }, status=status.HTTP_404_NOT_FOUND)

        return Response({
            "success": "True",
            "status_code": status.HTTP_200_OK,
            "message": "Asset List",
            "data": payload,
        }, status=status.HTTP_200_OK)




from django.db.models import Min

class SensorList(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request: HttpRequest):
        try:
            device_id = request.GET.get("device_id")

            # Get one unique sensor_id per device
            unique_ids = (
                SensorConfiguration.objects
                .filter(device_id=device_id)
                .values("sensor_id")
                .annotate(min_id=Min("id"))
                .values_list("min_id", flat=True)
            )

            sensors = SensorConfiguration.objects.filter(id__in=unique_ids).select_related("sensor")

            data = [
                {
                    "sensor__id": s.sensor.id,
                    "sensor__sensor_name": s.sensor.sensor_name,
                }
                for s in sensors
            ]

            return Response({
                "success": "True",
                "status_code": status.HTTP_200_OK,
                "message": "Sensor List",
                "data": data,
            })

        except Exception as e:
            response = f"on line {sys.exc_info()[-1].tb_lineno}, {str(e)}"
            return Response({"error": response}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MachineList(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            pk = request.data.get("asset_id")
            florida = timezone("Asia/Dhaka")
            florida_time = datetime.now(florida)
            time_stamp = florida_time.strftime("%Y-%m-%d %H:%M:%S")
            yesterday = datetime.today() - timedelta(days=1)
            before_date = yesterday.strftime("%Y-%m-%d %H:%M:%S")
            recieved_data = []

            device_list = Device.objects.filter(dev_asset=pk, dev_type=2)
            total_device = len(device_list)
            for device_data in device_list:
                devices_data = DeviceControlInfo.objects.filter(
                    dci_dev=device_data.id
                ).order_by("-id")[:1]
                if devices_data:
                    created_time = devices_data[0].dci_created_at
                    name = devices_data[0].dci_dev.dev_name
                    data = devices_data[0].dci_data
                    dev_id = devices_data[0].dci_dev_id
                    data_time = created_time.strftime("%d %b, %Y %I:%M %p")
                else:
                    created_time = ""
                    name = device_data.dev_name
                    data = "OFF"
                    dev_id = device_data.id
                    data_time = ""
                all_datas = {
                    "device_name": name,
                    "data": data,
                    "device_id": dev_id,
                    "time": data_time,
                }
                recieved_data.append(all_datas)
            asset_information = {
                "total_device": total_device,
                "device_info": recieved_data,
            }
            response = {
                "success": "True",
                "status code": status.HTTP_200_OK,
                "message": "Assets list for device control",
                "data": asset_information,
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except Exception as e:
            response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
            return Response(response, status=status.HTTP_201_CREATED)


class DeviceListWithData(APIView):
    permission_classes = (AllowAny,)

    def post(self, request):
        try:
            pk = request.data.get("asset_id")
            florida = timezone("Asia/Dhaka")
            florida_time = datetime.now(florida)
            time_stamp = florida_time.strftime("%Y-%m-%d %H:%M:%S")
            yesterday = datetime.today() - timedelta(days=1)
            before_date = yesterday.strftime("%Y-%m-%d %H:%M:%S")

            assets_info = AssetsProperties.objects.filter(id=pk).first()
            if assets_info:
                asset_data = []
                recieved_ph_data = []
                recieved_tds_data = []
                recieved_temp_data = []
                recieved_time_data = []

                device_list = Device.objects.filter(
                    dev_asset=assets_info.id, dev_type=1
                )
                for device_data in device_list:
                    devices_data = DeviceData.objects.filter(
                        dvd_dev=device_data.id,
                        dvd_created_at__range=[before_date, time_stamp],
                    ).order_by("-id")
                    for devices_datas in devices_data:
                        # all_ph_datas = [devices_datas.dvd_ph, 'tds': devices_datas.dvd_tds, 'temp': devices_datas.dvd_temp}
                        all_ph_datas = devices_datas.dvd_ph
                        recieved_ph_data.append(all_ph_datas)

                        all_tds_datas = devices_datas.dvd_tds
                        recieved_tds_data.append(all_tds_datas)

                        all_temp_datas = devices_datas.dvd_temp
                        recieved_temp_data.append(all_temp_datas)

                        create_date = devices_datas.dvd_created_at
                        all_time_datas = create_date.strftime("%H.%M")
                        recieved_time_data.append(all_time_datas)

                    total_device = len(device_list)
                    if device_data.dev_status == 0:
                        status_info = "Offline"
                    elif device_data.dev_status == 1:
                        status_info = "Online"
                    else:
                        status_info = "Problem"
                    device_wise_data_information = {
                        "id": device_data.id,
                        "status": status_info,
                        "total_device": total_device,
                        "name": device_data.dev_name,
                        "dev_description": device_data.dev_description,
                        "temp": recieved_temp_data,
                        "tds": recieved_tds_data,
                        "time": recieved_time_data,
                        "ph": recieved_ph_data,
                    }
                    asset_data.append(device_wise_data_information)

                # print(recieved_ph_data)
            # print(asset_data)
            response = {
                "success": "True",
                "status code": status.HTTP_200_OK,
                "message": "Assets wise data list",
                "data": asset_data,
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except Exception as e:
            response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
            return Response(response, status=status.HTTP_201_CREATED)


class DataHistoryPondWise(APIView):
    permission_classes = (AllowAny,)

    def get(self, request, pond):
        try:
            final_data = []
            devices = Device.objects.filter(dev_asset_id=pond, dev_type=1).values_list(
                "id"
            )
            for idx in devices:
                pond_wise_data = []
                print("line 639", idx[0])
                get_historical_device_data = (
                    DeviceData.objects.filter(dvd_dev_id=idx[0])
                    .order_by("-id")
                    .values()
                )
                for idy in get_historical_device_data:
                    # print("line 644", idy)
                    pond_wise_data.append(idy)
                final_data.append(pond_wise_data)
            response = {
                "success": "True",
                "status code": status.HTTP_200_OK,
                "message": "Pond wise Device Data History",
                "data": final_data,
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except Exception as e:
            response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
            return Response(response, status=status.HTTP_201_CREATED)


class DeviceDataSet(APIView):
    permission_classes = (AllowAny,)

    def get(self, request: HttpRequest):
        try:
            asset_id = request.GET.get('assst_id')
            sensor_id = request.GET.get("sensor_id")
            result_devices = list()
            company = request.GET.get("company_id",None)
            if company == None:
                company = request.user.company_id
            devices = Device.objects.filter(dev_asset_id=asset_id, company_id=company)

            
            switcher = {
                "daily": 1,
                "weekly": 7,
                "monthly": 30,
                "half-yearly": 180,
                "yearly": 365,
            }
            if request.GET.get("type") == "daily" or request.GET.get("type")==None:
                result_devices = pond_wise_devices_daily_data(
                    asset_id=asset_id,request=None, user=request.user,company=company,sensor_id=sensor_id
                )
            elif request.GET.get("type") == "weekly":
                result_devices = pond_wise_devices_weekly_data(
                    asset_id=asset_id, type=request.GET.get("type"), user=request.user, company=company,sensor_id=sensor_id
                )
            elif request.GET.get("type") == "monthly":
                result_devices = pond_wise_devices_monthly_data(
                    asset_id=asset_id,date_range= switcher.get(request.GET.get("type")), user=request.user, company=company,sensor_id=sensor_id
                )
            elif (
                request.GET.get("type") == "half-yearly"
                or request.GET.get("type") == "yearly"
            ):
                result_devices = pond_wise_devices_yearly_data(
                    asset_id=asset_id, type=request.GET.get("type"), user=request.user,company=company,sensor_id=sensor_id
                )

            response = {
                "success": "True",
                "status_code": status.HTTP_200_OK,
                "message": "Pond wise Device Data",
                "data": result_devices,
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except Exception as e:
            response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
            return Response(response, status=status.HTTP_201_CREATED)


class DeviceIndividualSensorData(APIView):
    permission_classes = (AllowAny,)

    def get(self, request, pond):
        try:
            devices = Device.objects.filter(dev_asset_id=pond, dev_type=1)
            result_devices = list()
            switcher = {
                "daily": 1,
                "weekly": 7,
                "monthly": 30,
                "half-yearly": 180,
                "yearly": 365,
            }
            if request.GET.get("type") == "daily":
                result_devices = pond_wise_devices_sensors_daily_data(devices)
            elif request.GET.get("type") == "weekly":
                result_devices = pond_wise_devices_sensors_weekly_data(
                    devices, request.GET.get("type")
                )
            elif request.GET.get("type") == "monthly":
                result_devices = pond_wise_devices_sensor_monthly_data(
                    devices, switcher.get(request.GET.get("type"))
                )
            elif (
                request.GET.get("type") == "half-yearly"
                or request.GET.get("type") == "yearly"
            ):
                result_devices = pond_wise_devices_sensor_yearly_data(
                    devices, request.GET.get("type")
                )

            response = {
                "success": "True",
                "status code": status.HTTP_200_OK,
                "message": "Pond wise Device Data",
                "data": result_devices,
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except Exception as e:
            response = "on line {}".format(sys.exc_info()[-1].tb_lineno), str(e)
            return Response(response, status=status.HTTP_201_CREATED)


class CameraListWithAssets(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        assets_data = []
        user_id = request.data.get("user_id")
        user_info = User.objects.filter(id=user_id, is_superuser=1)
        if user_info:
            asset_info = AssetsProperties.objects.all()
            serializer = AssetsCameraSerializer(asset_info, many=True)
            assets_data = serializer.data
        else:
            asset_info = AssetsProperties.ast_user.through.objects.filter(
                user_id=user_id
            )
            for asset_data in asset_info:
                camera_info = Camera.objects.filter(
                    cam_assets_id=asset_data.assetsproperties_id
                ).values()
                data = {
                    "id": asset_data.assetsproperties_id,
                    "ast_name": asset_data.assetsproperties.ast_name,
                    "camera_asset": camera_info,
                }
                assets_data.append(data)
        response = {
            "success": "True",
            "status code": status.HTTP_200_OK,
            "message": "Pond wise Camera",
            "data": assets_data,
        }
        return Response(response, status=status.HTTP_201_CREATED)


class MachineControl(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        florida = timezone("Asia/Dhaka")
        florida_time = datetime.now(florida)
        time_stamp = florida_time.strftime("%Y-%m-%d %H:%M:%S")
        yesterday = datetime.now(florida) + timedelta(seconds=10)
        after_time = yesterday.strftime("%Y-%m-%d %H:%M:%S")
        print(time_stamp, after_time)
        device_id = request.data.get("device_id")
        msg = request.data.get("message")
        device_info = Device.objects.filter(id=device_id).first()
        if device_info:
            device_serial = device_info.dev_serial_no
            phone_number = device_info.dev_sim_no

            if msg == "ON":
                value = "ON"
                publisher.client.connect(publisher.broker_address, port=publisher.port)
                publisher.client.publish(
                    "DMA/Fish_Farm/Padma_Shrimps/" + device_serial, value
                )

                time.sleep(2)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="ON",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is ON",
                    }
                    return Response(response)
                else:
                    pass
                time.sleep(2)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="ON",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is ON",
                    }
                    return Response(response)
                else:
                    pass
                time.sleep(4)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="ON",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is ON",
                    }
                    return Response(response)
                else:
                    pass
                publisher.client.connect(publisher.broker_address, port=publisher.port)
                publisher.client.publish(
                    "DMA/Fish_Farm/Padma_Shrimps/" + device_serial, value
                )
                time.sleep(2)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="ON",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is ON",
                    }
                    return Response(response)
                else:
                    pass
                time.sleep(3)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="ON",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is ON",
                    }
                    return Response(response)
                else:
                    pass

                url = "http://dma.com.bd:8888/send/sms"
                message = {"number": phone_number, "message": "ON"}
                r = requests.post(url, message)
                print(r)
                time.sleep(40)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="ON",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is ON",
                    }
                    return Response(response)
                else:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Internet Problem",
                    }
                    return Response(response)
            if msg == "OFF":
                value = "OFF"
                publisher.client.connect(publisher.broker_address, port=publisher.port)
                publisher.client.publish(
                    "DMA/Fish_Farm/Padma_Shrimps/" + device_serial, value
                )
                time.sleep(2)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="OFF",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is OFF",
                    }
                    return Response(response)
                else:
                    pass
                time.sleep(2)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="OFF",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is OFF",
                    }
                    return Response(response)
                else:
                    pass
                time.sleep(4)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="OFF",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is OFF",
                    }
                    return Response(response)
                else:
                    pass
                publisher.client.connect(publisher.broker_address, port=publisher.port)
                publisher.client.publish(
                    "DMA/Fish_Farm/Padma_Shrimps/" + device_serial, value
                )
                time.sleep(2)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="OFF",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is OFF",
                    }
                    return Response(response)
                else:
                    pass
                time.sleep(3)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="OFF",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is OFF",
                    }
                    return Response(response)
                else:
                    pass
                url = "http://dma.com.bd:8888/send/sms"
                message = {"number": phone_number, "message": "ON"}
                r = requests.post(url, message)
                print(r)
                time.sleep(40)
                devConInfo = DeviceControlInfo.objects.filter(
                    dci_dev=device_id,
                    dci_data="OFF",
                    dci_created_at__range=[time_stamp, after_time],
                ).order_by("-id")[:1]
                if devConInfo:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Device is OFF",
                    }
                    return Response(response)
                else:
                    response = {
                        "success": "True",
                        "status code": status.HTTP_200_OK,
                        "message": "Internet Problem",
                    }
                    return Response(response)
        else:
            response = {
                "success": "True",
                "status code": status.HTTP_200_OK,
                "message": "Device Not Found",
            }
            return Response(response)


# weather report
class WeatherReportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request:HttpRequest):
        
        company_id = request.GET.get("company_id",None)

        if company_id == None:
            asset_location = AssetsProperties.objects.select_related('district').filter(ast_user=request.user.id,district__isnull=False).first()
        else:
            asset_location = AssetsProperties.objects.select_related('district').filter(company_id=company_id).first()
        print(asset_location.district.district)
        weather = Weather.objects.filter(weather_city=asset_location.district.district).last()
        weather_data = WeatherSerializer(weather, many=False)
        print(weather)
        if weather:
            response = {
                "success": "True",
                "status_code": status.HTTP_200_OK,
                "data": weather_data.data,
            }

        else:
            response = {
                "success": "False",
                "status_code": status.HTTP_404_NOT_FOUND,
                "data": "No data found",
            }

        return Response(response)

class DeviceListAPIView(APIView):
    def get(self, request, format=None):
        company_id = request.GET.get('company_id')
        
        if company_id:
            devices = Device.objects.filter(company_id=company_id)
        else:
            devices = Device.objects.all()
        
        data = [{'id': device.id, 'dev_name': device.dev_name} for device in devices]
        return Response(data, status=status.HTTP_200_OK)
    
class UserManualDataAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        dhaka = timezone('Asia/Dhaka')
        dhaka_time = datetime.now(dhaka)
        time_stamp = dhaka_time.strftime("%Y-%m-%d %H:%M")
        print(time_stamp)

        user_id = request.data.get('user_id')
        asset_id = request.data.get('asset_id')
        man_ph = request.data.get('man_ph')
        man_ammonia = request.data.get('man_ammonia')
        man_DO = request.data.get('man_DO')
        man_tds = request.data.get('man_tds')
        man_temperature = request.data.get('man_temperature')
        man_updated_at_str = request.data.get('man_updated_at')  # Get the 'man_updated_at' value as a string

        try:
            man_updated_at = datetime.strptime(man_updated_at_str, "%Y-%m-%d %H:%M")  # Convert string to datetime
        except ValueError:
            return Response({
                "message": "Invalid 'man_updated_at' format. Use 'YYYY-MM-DD HH:MM'.",
                "success": False,
                "status_code": status.HTTP_400_BAD_REQUEST
            })

        serializer = UserManualDataSerializer(data={
            'man_assets': asset_id,
            'man_created_by': user_id,
            'man_ph': man_ph,
            'man_ammonia': man_ammonia,
            'man_DO': man_DO,
            'man_tds': man_tds,
            'man_temperature': man_temperature,
            'man_created_at': time_stamp,
            'man_updated_at': man_updated_at,  # Save the value of man_updated_at
        })

        try:
            serializer.is_valid(raise_exception=True)
            serializer.save()
            response = {
                'success': 'True',
                'status_code': status.HTTP_201_CREATED,
                'message': 'Data inserted successfully!'
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except:
            traceback.print_exc()
            return Response({
                "message": "Something went wrong. Check logs.",
                "success": False,
                "status_code": status.HTTP_500_INTERNAL_SERVER_ERROR
            })
   
        
class UserManualDataListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        user = request.user
        company_id = request.query_params.get('company_id')
        if company_id:
            try:
                company_id = int(company_id)
            except ValueError:
                return Response({"message": "Invalid company ID. Must be an integer."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            company_id = user.company.id

        asset_id = request.query_params.get('asset_id')
        if asset_id:
            try:
                asset_id = int(asset_id)
            except ValueError:
                return Response({"message": "Invalid asset ID. Must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                asset = AssetsProperties.objects.get(pk=asset_id, company_id=company_id)
            except AssetsProperties.DoesNotExist:
                return Response({"message": "Asset not found or does not belong to the specified company."},
                                status=status.HTTP_404_NOT_FOUND)

            manual_data = UserManualData.objects.filter(man_assets=asset)
        else:
            manual_data = UserManualData.objects.filter(man_assets__company_id=company_id)

        data = [{
            'id': item.id,
            'man_assets': item.man_assets.id,
            'asset_name': item.man_assets.ast_name,
            'man_ph': item.man_ph,
            'man_ammonia': item.man_ammonia,
            'man_DO': item.man_DO,
            'man_tds': item.man_tds,
            'man_temperature': item.man_temperature,
            'man_created_by': item.man_created_by.first_name + ' ' + item.man_created_by.last_name,
            'datetime': (
                item.man_updated_at.strftime('%Y-%m-%d %I:%M %p') 
                if item.man_updated_at is not None
                else item.man_created_at.strftime('%Y-%m-%d %I:%M %p')
            ),
        } for item in manual_data]

        data = sorted(data, key=lambda x: x['datetime'], reverse=True)
                
        if not data:
            return Response({"success": False, "status_code": 404, "data": []}, status=status.HTTP_404_NOT_FOUND)

        return Response({"success": True, "status_code": 200, "data": data}, status=status.HTTP_200_OK)


class GenerateCSVAPIView(APIView):
    permission_classes = [IsAuthenticated]  
    
    def generate_excel_data(self, asset, start_date=None, end_date=None):
        
        manual_data = UserManualData.objects.filter(man_assets=asset)

        if start_date and end_date:
            manual_data = manual_data.filter(
                Q(man_updated_at__date__range=[start_date, end_date], man_updated_at__isnull=False) |
                Q(man_created_at__date__range=[start_date, end_date])
            )
        data_list = []
        
        for manual_item in manual_data:
            if manual_item.man_updated_at is not None:
                man_datetime = manual_item.man_updated_at
            else:
                man_datetime = manual_item.man_created_at
            
            print("Manual date time: ",man_datetime)
            print("created at gte: ",man_datetime - timedelta(minutes=120))
            print("created at lte: ",man_datetime + timedelta(minutes=120))
            
            device_data_within_range = DeviceData.objects.filter(
                dvd_created_at__gte=man_datetime - timedelta(minutes=120),
                dvd_created_at__lte=man_datetime + timedelta(minutes=120),
                dvd_dev__dev_asset=asset
            ).order_by('dvd_created_at')
            
            for data in device_data_within_range:
                print("Device data \n")
                print(data.dvd_created_at)
            if device_data_within_range.exists():
                closest_device_data = device_data_within_range.first()
                print("Closest Device Data:", closest_device_data.dvd_created_at)
                             
                delta_ammonia = (
                    float(closest_device_data.dvd_ammonia) - float(manual_item.man_ammonia)
                ) if closest_device_data.dvd_ammonia and manual_item.man_ammonia else ""

                delta_DO = (
                    float(closest_device_data.dvd_do) - float(manual_item.man_DO)
                ) if closest_device_data.dvd_do and manual_item.man_DO else ""

                delta_temperature = (
                    float(closest_device_data.dvd_temp) - float(manual_item.man_temperature)
                ) if closest_device_data.dvd_temp and manual_item.man_temperature else ""

                delta_tds = (
                    float(closest_device_data.dvd_tds) - float(manual_item.man_tds)
                ) if closest_device_data.dvd_tds and manual_item.man_tds else ""

                delta_ph = (
                    float(closest_device_data.dvd_ph) - float(manual_item.man_ph)
                ) if closest_device_data.dvd_ph and manual_item.man_ph else ""
                
                man_updated_at_12hr = man_datetime.strftime('%Y-%m-%d %H:%M:%S %p')
                man_created_at_12hr = manual_item.man_created_at.strftime('%Y-%m-%d %H:%M:%S %p')
                device_data_time_12hr = closest_device_data.dvd_created_at.strftime('%Y-%m-%d %H:%M:%S %p')
                print("closest device data formatted time: ",closest_device_data.dvd_created_at)

                variance_data = {
                    'asset_name': asset.ast_name,
                    'man_updated_at': man_updated_at_12hr,  # Format the date
                    'man_created_at': man_created_at_12hr, # Format the date
                    'device_data_time': device_data_time_12hr,
                    'man_ph': manual_item.man_ph,
                    'man_ammonia': manual_item.man_ammonia,
                    'man_DO': manual_item.man_DO,
                    'man_tds': manual_item.man_tds,
                    'man_temperature': manual_item.man_temperature,
                    'device_ph': closest_device_data.dvd_ph,
                    'device_ammonia': closest_device_data.dvd_ammonia,
                    'device_DO': closest_device_data.dvd_do,
                    'device_tds': closest_device_data.dvd_tds,
                    'device_temperature': closest_device_data.dvd_temp,
                    'delta_ph': delta_ph,
                    'delta_ammonia': delta_ammonia,
                    'delta_DO': delta_DO,
                    'delta_tds': delta_tds,
                    'delta_temperature': delta_temperature,

                }
                print(variance_data)
                data_list.append(variance_data)
        return data_list

    def post(self, request, format=None):
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')
        asset_id = request.data.get('asset_id')
        company_id = request.data.get('company_id')
        try:
            if start_date_str:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            else:
                start_date = None

            if end_date_str:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            else:
                end_date = None
        except ValueError:
            return Response({"success": False, "message": "Invalid date format. Use 'YYYY-MM-DD'."}, status=status.HTTP_400_BAD_REQUEST)
        if company_id:
            company = get_object_or_404(Company, pk=company_id)
        else:
            company = request.user.company
        if asset_id:
            try:
                asset = AssetsProperties.objects.get(pk=asset_id, company=company)
                data_list = self.generate_excel_data(asset, start_date, end_date)
                file_name = f"{uuid.uuid4().hex}.xlsx"
            except AssetsProperties.DoesNotExist:
                return Response({"success": False, "status_code": 404, "message": "Asset not found or does not belong to your company."}, status=status.HTTP_404_NOT_FOUND)
        else:
            data_list = []
            all_assets = AssetsProperties.objects.filter(company=company)
            for asset in all_assets:
                data_list.extend(self.generate_excel_data(asset, start_date, end_date))
            file_name = f"{uuid.uuid4().hex}.xlsx"

        file_path = os.path.join(settings.BASE_DIR, file_name)

        # Create an Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Additional rows for the Excel file
        additional_rows = [
            ["", "MoreFish Report", "", ""],
            ["", f"Company name = {company}", "", "", f"User = {request.user.first_name} {request.user.last_name}", "", ""],
            ["", "", "", ""],
        ]

        # Write the additional rows to the worksheet
        for row in additional_rows:
            worksheet.append(row)

        # Header row with styling
        header_row = [
            "Pond Name",
            "Manual Data Input Time",
            "Manual Data Registered Time",
            "Device Data Time",
            "",
            "Manual pH",
            "Manual Ammonia",
            "Manual DO",
            "Manual TDS",
            "Manual Temperature",
            "",
            "Device pH",
            "Device Ammonia",
            "Device DO",
            "Device TDS",
            "Device Temperature",
            "",
            "Delta pH",
            "Delta Ammonia",
            "Delta DO",
            "Delta TDS",
            "Delta Temperature",
        ]
        worksheet.append(header_row)

        # Write data rows
        for item in data_list:
            # Check if the delta values are numeric before rounding
            delta_ammonia = round(item['delta_ammonia'], 2) if isinstance(item['delta_ammonia'], (int, float)) else item['delta_ammonia']
            delta_DO = round(item['delta_DO'], 2) if isinstance(item['delta_DO'], (int, float)) else item['delta_DO']
            delta_temperature = round(item['delta_temperature'], 2) if isinstance(item['delta_temperature'], (int, float)) else item['delta_temperature']
            delta_tds = round(item['delta_tds'], 2) if isinstance(item['delta_tds'], (int, float)) else item['delta_tds']
            delta_ph = round(item['delta_ph'], 2) if isinstance(item['delta_ph'], (int, float)) else item['delta_ph']

            data_row = [
                item['asset_name'],
                item['man_updated_at'],
                item['man_created_at'],
                item['device_data_time'],
                '',
                item['man_ph'],
                item['man_ammonia'],
                item['man_DO'],
                item['man_tds'],
                item['man_temperature'],
                '',
                item['device_ph'],
                item['device_ammonia'],
                item['device_DO'],
                item['device_tds'],
                item['device_temperature'],
                '',
                delta_ph,
                delta_ammonia,
                delta_DO,
                delta_tds,
                delta_temperature,
            ]

            worksheet.append(data_row)

        # Save the Excel file
        workbook.save(file_path)

        download_link = request.build_absolute_uri(reverse('csv-download', args=[file_name]))
        if not data_list:
            return JsonResponse({"success": False, "status_code": 204, "message": "No data found for the provided filters.", "download_link": download_link}, status=status.HTTP_204_NO_CONTENT)

        return JsonResponse({"success": True, "status_code": 200, "message": "Excel file generated successfully.", "download_link": download_link}, status=status.HTTP_200_OK)


class DownloadCSVView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, file_name, format=None):
        file_path = os.path.join(settings.BASE_DIR, file_name)
        if not os.path.exists(file_path):
            return JsonResponse({"success": False, "status_code": 404, "message": "File not found."}, status=status.HTTP_404_NOT_FOUND)

        with open(file_path, 'rb') as file:
            response = HttpResponse(file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Set the Content-Disposition header to rename the downloaded file
        response['Content-Disposition'] = f'attachment; filename="MoreFish_report.xlsx"'

        # Delete the generated Excel file after it has been downloaded
        os.remove(file_path)

        return response



class GetDevicesForCompanyView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        if self.request.is_ajax():
            company_id = self.request.GET.get('company_id')
            devices = Device.objects.filter(company_id=company_id).values('id', 'dev_name')
            return JsonResponse({'devices': list(devices)})

        return JsonResponse({})

    
class ComplainCategoryView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        get_complain_category = ComplainCategory.objects.all().values()
        complain_list = []
        for idx in get_complain_category:
            complain_list.append(idx['category_name'])
        if get_complain_category:
            response = {
                'success': 'True',
                'status_code': status.HTTP_200_OK,
                'data': complain_list
            }
        else:
            response = {
                'success': 'False',
                'status_code': status.HTTP_404_NOT_FOUND,
                'data': complain_list
            }
        return Response(response)


class ComplainSubmitView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # user_id = request.data.get('company_id')
        complain_category = request.data.get('complain_category')
        asset_id = request.POST.get('asset_id')
        complain_title = request.data.get('complain_title')
        complain_desc = request.data.get('complain_desc')
        complain_img = request.data.get('complain_img')
        print("line 1404", type(complain_img))
        
        assets_properties = get_object_or_404(AssetsProperties, id=asset_id)        
        submit_complain = Complain.objects.create(
            user_id=request.user.id,
            complain_asset=assets_properties,
            complain_title=complain_title,
            complain_description=complain_desc,
            complain_image=complain_img,
            complain_category=complain_category
        )
        print(submit_complain)
        if submit_complain:
            response = {
                'success': 'True',
                'status_code': status.HTTP_201_CREATED,
                'msg': "Complain Created Successfully"
            }
        else:
            response = {
                'success': 'False',
                'status_code': status.HTTP_400_BAD_REQUEST,
                'msg': "Complain Not Created"
            }
        return Response(response)


class ComplainListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        requesting_user = request.user
        requesting_user_company = requesting_user.company
        # users = User.objects.filter(company=requesting_user_company)
        get_complain = Complain.objects.filter(user__company=request.user.company).values()
        print(requesting_user_company)
        # print(users)
        if get_complain:
            response = {
                'success': 'True',
                'status_code': status.HTTP_200_OK,
                'data': get_complain
            }
        else:
            response = {
                'success': 'False',
                'status_code': status.HTTP_404_NOT_FOUND,
                'data': get_complain
            }

        return Response(response)
  


class GenerateAssetDataExcel(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')
        company_id = request.data.get('company_id')
        asset_id = request.data.get('asset_id') 

        try:
            if start_date_str:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            else:
                start_date = None

            if end_date_str:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            else:
                end_date = None

            if start_date and end_date and start_date > end_date:
                return JsonResponse({
                    'success': 'False',
                    'status_code': status.HTTP_400_BAD_REQUEST,
                    'msg': "Start date cannot be greater than end date."
                }, status=status.HTTP_400_BAD_REQUEST)

        except ValueError:
            return JsonResponse({
                'success': 'False',
                'status_code': status.HTTP_400_BAD_REQUEST,
                'msg': "Invalid date format. Use 'YYYY-MM-DD'."
            }, status=status.HTTP_400_BAD_REQUEST)

        company = get_object_or_404(Company, pk=company_id) if company_id else None

        # Fetch all assets for the given company
        assets = AssetsProperties.objects.filter(company=company_id) if company_id else AssetsProperties.objects.all()

        device_data = DeviceData.objects.all()

        if start_date and end_date:
            device_data = device_data.filter(dvd_created_at__date__range=[start_date, end_date])

        if company:
            device_data = device_data.filter(company=company)

        if asset_id:
            device_data = device_data.filter(dvd_dev__dev_asset=asset_id)

        device_data = device_data.order_by('dvd_created_at')

        if not device_data.exists():
            return JsonResponse({
                'success': 'False',
                'status_code': status.HTTP_204_NO_CONTENT,
                'msg': "No data found for the provided filters."
            }, status=status.HTTP_204_NO_CONTENT)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        additional_rows = [
            ["", "MoreFish Device Data Report", "", ""],
            ["", f"Company name = {company}", "", "", f"User = {request.user.first_name} {request.user.last_name}", "", ""],
            ["", "", "", ""],
        ]

        for row in additional_rows:
            worksheet.append(row)
           
        header = [
            "Asset name",
            "Created At",
            "pH",
            "TDS",
            "Temperature",
            "Ammonia",
            "DO",
            "Alkalinity",
            "Hardness"
        ]

        worksheet.append(header)

        for asset in assets:
            asset_name = asset.ast_name

            asset_device_data = device_data.filter(dvd_dev__dev_asset=asset.pk)

            for item in asset_device_data:
                created_at = item.dvd_created_at.replace(tzinfo=None)
                data_row = [
                    asset_name,
                    created_at,
                    item.dvd_ph,
                    item.dvd_tds,
                    item.dvd_temp,
                    item.dvd_ammonia,
                    item.dvd_do,
                    item.dvd_alkalinity,
                    item.dvd_hardness,
                ]
                worksheet.append(data_row)

        file_name = f"{uuid.uuid4().hex}.xlsx"
        file_path = os.path.join(settings.BASE_DIR, file_name)

        workbook.save(file_path)

        download_link = request.build_absolute_uri(reverse('download-excel-assetwisedata', args=[start_date,end_date,file_name]))

        return JsonResponse({
            'success': 'True',
            'status_code': status.HTTP_200_OK,
            'msg': "Excel file generated successfully.",
            'download_link': download_link
        }, status=status.HTTP_200_OK)

        
class DownloadExcelAssetDataView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, file_name, start_date, end_date, format=None):
        file_path = os.path.join(settings.BASE_DIR, file_name)
        if not os.path.exists(file_path):
            return JsonResponse({"message": "File not found."}, status=status.HTTP_404_NOT_FOUND)

        with open(file_path, 'rb') as file:
            response = HttpResponse(file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Asset Wise Report {datetime.now()}.xlsx"'

        os.remove(file_path)

        return response


# views.py
import logging
from django.http import JsonResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from device.models import Aerator
from .serializers import AeratorCommandSerializer, AeratorInfoSerializer  # adjust path


class AeratorCommandAPI(APIView):
    """
    POST /devices/aerators/command/
    {
      "aerator_id": "A001",
      "command": 1   # 1 = ON, 0 = OFF
    }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        logging.info(f"[CMD] POST data: {request.data}")

        serializer = AeratorCommandSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        aerator_id = serializer.validated_data['aerator_id']
        cmd = serializer.validated_data['command']  # 0 or 1
        turn_on = (cmd == 1)

        try:
            aer = Aerator.objects.get(aerator_id=aerator_id)
        except Aerator.DoesNotExist:
            return JsonResponse(
                {
                    'success': 'False',
                    'status_code': status.HTTP_404_NOT_FOUND,
                    'msg': 'Aerator not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        if not aer.is_active:
            return JsonResponse(
                {
                    'success': 'False',
                    'status_code': status.HTTP_400_BAD_REQUEST,
                    'msg': 'Aerator inactive.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            cmd_log = publish_command(aer, turn_on)
        except Exception as e:
            return JsonResponse(
                {
                    'success': 'False',
                    'status_code': status.HTTP_502_BAD_GATEWAY,
                    'aerator_id': aerator_id,
                    'command': cmd,
                    'msg': f'Failed to publish command: {e}'
                },
                status=status.HTTP_502_BAD_GATEWAY
            )

        return JsonResponse(
            {
                'success': 'True',
                'status_code': status.HTTP_200_OK,
                'aerator_id': aerator_id,
                'command': cmd,
                'command_log_id': cmd_log.id,
                'msg': 'Command sent successfully.'
            },
            status=status.HTTP_200_OK
        )
